import openpyxl
from selenium import webdriver;
from selenium.webdriver.common.keys import Keys; 
from selenium.webdriver.support.ui import Select
import time
import pandas as pd
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.remote.webelement import WebElement
import unittest
import logging
from random import randint
from selenium.common.exceptions import NoSuchElementException

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)


def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

driver = webdriver.Chrome(executable_path='./drivers/chromedriver')

driver.get("https://athena-ci.herokuapp.com")
driver.maximize_window()
time.sleep(2)

path="C://Users//Kal-el//Desktop//selenium_help-main - Copy//data.xlsx"

a = ActionChains(driver)
m = driver.find_element_by_xpath("//a[@class='active']")
a.move_to_element(m).perform()
time.sleep(0.3)

n = driver.find_element_by_xpath("//a[normalize-space()='About']")
a.move_to_element(n).perform()
time.sleep(0.3)

o = driver.find_element_by_xpath("//span[normalize-space()='Access']")
a.move_to_element(o).perform()
time.sleep(0.3)

p = driver.find_element_by_xpath("//a[normalize-space()='User Sign up']")
a.move_to_element(p).perform()
time.sleep(0.3)

q = driver.find_element_by_xpath("//a[normalize-space()='User Login']")
a.move_to_element(q).perform()
time.sleep(0.3)

submit = driver.find_element_by_xpath("//a[normalize-space()='User Login']")
submit.click()

def slow_type(element: WebElement, text: str, delay: float=0.1):
    """Send a text to an element one character at a time with a delay."""
    for character in text:
        element.send_keys(character)
        time.sleep(delay)

username = driver.find_element_by_xpath("//input[@placeholder='Enter Email ID']")        
text = "vinay.gat1@gmail.com"
slow_type(username, text)       

password = driver.find_element_by_xpath("//input[@placeholder='Enter Password']")
text = "superman"
slow_type(password, text)

time.sleep(0.5)

submit = driver.find_element_by_xpath("//input[@value='login']")
submit.click()

time.sleep(1)

alert_obj = driver.switch_to.alert
alert_obj.accept()
 
r = driver.find_element_by_xpath("//a[normalize-space()='HOME']")
a.move_to_element(r).perform()
time.sleep(0.3)

s = driver.find_element_by_xpath("//a[normalize-space()='UPLOAD NOTES']")
a.move_to_element(s).perform() 
time.sleep(0.3)

submit = driver.find_element_by_xpath("//a[normalize-space()='UPLOAD NOTES']")
submit.click()

rows=getRowCount(path, 'Sheet1')

for r in range(2,rows+1):
    branch=readData(path,"Sheet1",r,1)
    subject=readData(path,"Sheet1",r,2)
    file=readData(path,"Sheet1",r,3)
    type=readData(path,"Sheet1",r,4)
    desc=readData(path,"Sheet1",r,5)

    
    status_select = Select(driver.find_element_by_xpath("//select[@name='branch']"))
    status_select.select_by_value(branch)
    time.sleep(1)
    subject_s = driver.find_element_by_xpath("//input[@placeholder='Enter Subject']")        
    text = (subject)
    slow_type(subject_s, text)
    time.sleep(1)
    add1 = driver.find_element_by_xpath("//input[@name='notesfile']")  
    add1.send_keys(file)
    time.sleep(1)
    status_type = Select(driver.find_element_by_xpath("//select[@name='filetype']"))
    status_type.select_by_value(type)
    time.sleep(1)
    desc_p = driver.find_element_by_xpath("//textarea[@name='description']")        
    text = (desc)
    slow_type(desc_p, text)
    time.sleep(1)
    driver.find_element_by_xpath("//input[@value='Submit']").click()

    alert_obj = driver.switch_to.alert
    alert_obj.accept()

    # driver.implicitly_wait(10)
    # if driver.title=="Notes Sharing Site":
    #     print("test is passed")
    #     writeData(path,"sheet1",r,6,"test passed")    
    # else:    
    #     print("test failed")
    #     writeData(path,"sheet1",r,6,"test failed")

    driver.implicitly_wait(5) # seconds
    driver.get("https://athena-ci.herokuapp.com/view_mynotes")
    myDynamicElement = driver.find_element_by_xpath("//a[normalize-space()='UPLOAD NOTES']").click()
    
   