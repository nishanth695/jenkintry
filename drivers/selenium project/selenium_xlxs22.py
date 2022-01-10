import openpyxl
from selenium import webdriver;
from selenium.webdriver.common.keys import Keys; 
from selenium.webdriver.support.ui import Select
import time
import pandas as pd
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.remote.webelement import WebElement
import unittest
import logging
from random import randint
from selenium.common.exceptions import NoSuchElementException

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)


def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file) 
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

driver = webdriver.Chrome(executable_path='./drivers/chromedriver')

driver.get("https://athena-ci.herokuapp.com/contact")
driver.maximize_window()
time.sleep(2)

path="C://Users//Kal-el//Desktop//selenium_help-main - Copy//data2.xlsx"

rows=getRowCount(path, 'Sheet1')

for r in range(2,rows+1):
    branch=readData(path,"Sheet1",r,1)
    subject=readData(path,"Sheet1",r,2)
    file=readData(path,"Sheet1",r,3)
    type=readData(path,"Sheet1",r,4)
    desc=readData(path,"Sheet1",r,5)

    driver.find_element_by_xpath("//input[@placeholder='full name']").send_keys(branch)
    time.sleep(2)
    driver.find_element_by_xpath("//input[@placeholder='your email']").send_keys(subject)
    time.sleep(2)
    driver.find_element_by_xpath("//input[@placeholder='your contact number']").send_keys(file)
    time.sleep(2)
    driver.find_element_by_xpath("//input[@placeholder='Subject']").send_keys(type)
    time.sleep(2)
    driver.find_element_by_xpath("//textarea[@placeholder='Send Message...']").send_keys(desc)
    time.sleep(2)
    driver.find_element_by_xpath("//input[@value='send now']").click()


    alert_obj = driver.switch_to.alert
    alert_obj.accept()

if driver.title=="Online Note Sharing":
    print("test is passed")
else :
    print("test failed")

    driver.find_element_by_xpath("//a[@class='logo']").click